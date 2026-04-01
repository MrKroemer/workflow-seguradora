from decimal import Decimal

from openpyxl import Workbook

from rpa_corretora.integrations.segfy_gateway import SegfyGateway


def test_segfy_register_payment_writes_queue_when_no_api(tmp_path) -> None:
    queue_file = tmp_path / "segfy_queue.jsonl"
    gateway = SegfyGateway(queue_path=queue_file)

    success = gateway.register_payment(
        commitment_id="agenda-1",
        description="Baixa de parcela",
    )

    assert success is False
    assert queue_file.exists()
    content = queue_file.read_text(encoding="utf-8")
    assert "agenda-1" in content
    assert "Baixa de parcela" in content


def test_segfy_fetch_policy_data_from_export_xlsx(tmp_path) -> None:
    export_file = tmp_path / "segfy_export.xlsx"
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Polices"
    ws["A1"] = "Apolice"
    ws["B1"] = "Premio Total"
    ws["C1"] = "Comissao"
    ws["A2"] = "PB-2001"
    ws["B2"] = "1.250,00"
    ws["C2"] = "187,50"
    workbook.save(export_file)

    gateway = SegfyGateway(export_xlsx_path=export_file)
    rows = gateway.fetch_policy_data()

    assert len(rows) == 1
    assert rows[0].policy_id == "PB-2001"
    assert rows[0].premio_total == Decimal("1250.00")
    assert rows[0].comissao == Decimal("187.50")
