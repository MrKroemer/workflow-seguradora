from __future__ import annotations

from collections import defaultdict
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from urllib.parse import urljoin
from urllib.request import Request, urlopen
import json
import unicodedata

from openpyxl import load_workbook

from rpa_corretora.domain.models import CashflowEntry, FollowupRecord, PolicyRecord, SegfyPolicyData


def _to_date(value: object) -> date | None:
    if value is None:
        return None
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            from datetime import datetime as _dt
            return _dt.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _normalize(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text == "":
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    return text.strip().upper()


def _to_decimal(value: object) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip().replace(".", "").replace(",", ".")
    if text == "":
        return Decimal("0")
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return Decimal("0")

class SegfyGateway:
    def __init__(
        self,
        *,
        username: str | None = None,
        password: str | None = None,
        api_base_url: str | None = None,
        api_token: str | None = None,
        api_login_path: str = "/auth/login",
        api_policies_path: str = "/policies",
        api_register_payment_path: str = "/payments/register",
        export_xlsx_path: str | Path | None = None,
        queue_path: str | Path = "outputs/segfy_payment_queue.jsonl",
        timeout_seconds: int = 20,
        allow_queue_fallback: bool = True,
    ) -> None:
        self.username = (username or "").strip()
        self.password = (password or "").strip()
        self.api_base_url = (api_base_url or "").strip().rstrip("/")
        self.api_token = (api_token or "").strip()
        self.api_login_path = api_login_path
        self.api_policies_path = api_policies_path
        self.api_register_payment_path = api_register_payment_path
        self.export_xlsx_path = Path(export_xlsx_path) if export_xlsx_path else None
        self.queue_path = Path(queue_path)
        self.timeout_seconds = timeout_seconds
        self.allow_queue_fallback = allow_queue_fallback

    def fetch_policy_data(self) -> list[SegfyPolicyData]:
        api_data = self._fetch_policy_data_from_api()
        if api_data is not None:
            return api_data
        return self._fetch_policy_data_from_export()

    def fetch_full_policies(self) -> list[PolicyRecord]:
        """Lê todas as colunas disponíveis do export Segfy e retorna PolicyRecords completos."""
        return self._fetch_full_policies_from_export()

    def import_documents(self) -> int:
        return 0

    def sync_policies(self, policies: list[PolicyRecord]) -> int:
        if not self.api_base_url:
            return self._queue_bulk("sync_policies", [{"policy_id": p.policy_id, "insured_name": p.insured_name, "insurer": p.insurer, "vig": p.vig.isoformat(), "premio_total": str(p.premio_total), "comissao": str(p.comissao), "status_pgto": p.status_pgto, "vehicle_item": p.vehicle_item} for p in policies])
        token = self._auth_token()
        if token is None:
            return self._queue_bulk("sync_policies", [{"policy_id": p.policy_id} for p in policies])
        synced = 0
        for policy in policies:
            payload = {"policy_id": policy.policy_id, "insured_name": policy.insured_name, "insurer": policy.insurer, "vig": policy.vig.isoformat(), "premio_total": str(policy.premio_total), "comissao": str(policy.comissao), "status_pgto": policy.status_pgto, "vehicle_item": policy.vehicle_item, "sinistro_open": policy.sinistro_open, "endosso_open": policy.endosso_open}
            result = self._request_json(method="POST", path="/policies/sync", body=payload, token=token)
            if result is not None:
                synced += 1
        return synced

    def sync_followups(self, followups: list[FollowupRecord]) -> int:
        if not self.api_base_url:
            return self._queue_bulk("sync_followups", [{"insured_name": f.insured_name, "month": f.month, "fase": f.fase, "status": f.status, "renewal_kind": f.renewal_kind} for f in followups])
        token = self._auth_token()
        if token is None:
            return self._queue_bulk("sync_followups", [{"insured_name": f.insured_name} for f in followups])
        synced = 0
        for followup in followups:
            payload = {"insured_name": followup.insured_name, "month": followup.month, "fase": followup.fase, "status": followup.status, "renewal_kind": followup.renewal_kind}
            result = self._request_json(method="POST", path="/followups/sync", body=payload, token=token)
            if result is not None:
                synced += 1
        return synced

    def sync_cashflow(self, entries: list[CashflowEntry]) -> int:
        if not self.api_base_url:
            return self._queue_bulk("sync_cashflow", [{"date": e.date.isoformat(), "value": str(e.value), "insurer": e.insurer, "specification": e.specification} for e in entries])
        token = self._auth_token()
        if token is None:
            return self._queue_bulk("sync_cashflow", [{"date": e.date.isoformat()} for e in entries])
        synced = 0
        for entry in entries:
            payload = {"date": entry.date.isoformat(), "value": str(entry.value), "insurer": entry.insurer, "specification": entry.specification, "source": entry.source}
            result = self._request_json(method="POST", path="/cashflow/sync", body=payload, token=token)
            if result is not None:
                synced += 1
        return synced

    def register_incident(self, *, policy_id: str, incident_type: str, description: str) -> bool:
        payload = {"policy_id": policy_id, "incident_type": incident_type, "description": description}
        if not self.api_base_url:
            return self._queue_single("register_incident", payload)
        token = self._auth_token()
        if token is None:
            return self._queue_single("register_incident", payload)
        result = self._request_json(method="POST", path="/incidents/register", body=payload, token=token)
        return result is not None

    def update_commission_status(self, *, policy_id: str, status: str) -> bool:
        payload = {"policy_id": policy_id, "status": status}
        if not self.api_base_url:
            return self._queue_single("update_commission", payload)
        token = self._auth_token()
        if token is None:
            return self._queue_single("update_commission", payload)
        result = self._request_json(method="POST", path="/commissions/update", body=payload, token=token)
        return result is not None

    def register_renewal(self, *, policy_id: str, phase: str, status: str) -> bool:
        payload = {"policy_id": policy_id, "phase": phase, "status": status}
        if not self.api_base_url:
            return self._queue_single("register_renewal", payload)
        token = self._auth_token()
        if token is None:
            return self._queue_single("register_renewal", payload)
        result = self._request_json(method="POST", path="/renewals/register", body=payload, token=token)
        return result is not None

    def _queue_single(self, action: str, payload: dict[str, str]) -> bool:
        if not self.allow_queue_fallback:
            return False
        self.queue_path.parent.mkdir(parents=True, exist_ok=True)
        with self.queue_path.open("a", encoding="utf-8") as file:
            file.write(json.dumps({"action": action, **payload}) + "\n")
        return False

    def _queue_bulk(self, action: str, items: list[dict[str, object]]) -> int:
        if not self.allow_queue_fallback:
            return 0
        self.queue_path.parent.mkdir(parents=True, exist_ok=True)
        with self.queue_path.open("a", encoding="utf-8") as file:
            for item in items:
                file.write(json.dumps({"action": action, **item}) + "\n")
        return 0

    def register_payment(self, *, commitment_id: str, description: str) -> bool:
        payload = {
            "commitment_id": commitment_id,
            "description": description,
        }
        if self._post_payment(payload):
            return True

        if not self.allow_queue_fallback:
            return False

        self.queue_path.parent.mkdir(parents=True, exist_ok=True)
        with self.queue_path.open("a", encoding="utf-8") as file:
            file.write(json.dumps(payload) + "\n")
        return False

    def _auth_token(self) -> str | None:
        if self.api_token:
            return self.api_token
        if not self.api_base_url or not self.username or not self.password:
            return None

        payload = self._request_json(
            method="POST",
            path=self.api_login_path,
            body={"username": self.username, "password": self.password},
            token=None,
        )
        if not isinstance(payload, dict):
            return None
        candidates = (
            payload.get("access_token"),
            payload.get("token"),
            payload.get("jwt"),
        )
        for token in candidates:
            if isinstance(token, str) and token.strip():
                return token.strip()
        return None

    def _fetch_policy_data_from_api(self) -> list[SegfyPolicyData] | None:
        if not self.api_base_url:
            return None
        token = self._auth_token()
        if token is None:
            print("[Segfy] Sem token para consulta de apolices via API.")
            return None

        payload = self._request_json(
            method="GET",
            path=self.api_policies_path,
            token=token,
        )
        if payload is None:
            return None

        items: list[object]
        if isinstance(payload, list):
            items = payload
        elif isinstance(payload, dict):
            maybe_items = payload.get("items")
            if isinstance(maybe_items, list):
                items = maybe_items
            elif isinstance(payload.get("data"), list):
                items = payload["data"]  # type: ignore[index]
            else:
                items = []
        else:
            items = []

        parsed: list[SegfyPolicyData] = []
        for raw in items:
            if not isinstance(raw, dict):
                continue
            policy_id = str(
                raw.get("policy_id")
                or raw.get("id")
                or raw.get("numero_apolice")
                or raw.get("apolice")
                or ""
            ).strip()
            if not policy_id:
                continue
            premio = _to_decimal(raw.get("premio_total") or raw.get("premio") or raw.get("valor_premio"))
            comissao = _to_decimal(raw.get("comissao") or raw.get("valor_comissao"))
            parsed.append(
                SegfyPolicyData(
                    policy_id=policy_id,
                    premio_total=premio,
                    comissao=comissao,
                )
            )
        return parsed

    def _post_payment(self, payload: dict[str, str]) -> bool:
        if not self.api_base_url:
            return False
        token = self._auth_token()
        if token is None:
            return False
        response = self._request_json(
            method="POST",
            path=self.api_register_payment_path,
            body=payload,
            token=token,
        )
        return response is not None

    def _request_json(
        self,
        *,
        method: str,
        path: str,
        body: dict[str, object] | None = None,
        token: str | None,
    ) -> dict[str, object] | list[object] | None:
        if not self.api_base_url:
            return None
        url = urljoin(f"{self.api_base_url}/", path.lstrip("/"))
        encoded = None
        headers = {"Content-Type": "application/json"}
        if body is not None:
            encoded = json.dumps(body).encode("utf-8")
        if token:
            headers["Authorization"] = f"Bearer {token}"
        request = Request(url, data=encoded, headers=headers, method=method)
        try:
            with urlopen(request, timeout=self.timeout_seconds) as response:
                raw = response.read().decode("utf-8")
                if raw.strip() == "":
                    return {}
                return json.loads(raw)
        except Exception as exc:
            print(f"[Segfy] Falha em {method} {url}: {exc}")
            return None

    def _fetch_policy_data_from_export(self) -> list[SegfyPolicyData]:
        if self.export_xlsx_path is None or not self.export_xlsx_path.exists():
            return []

        workbook = load_workbook(self.export_xlsx_path, data_only=True)
        parsed: list[SegfyPolicyData] = []
        for ws in workbook.worksheets:
            header_row, header_map = self._find_headers(ws)
            if header_row is None:
                continue
            policy_col = header_map.get("POLICY")
            premio_col = header_map.get("PREMIO")
            comissao_col = header_map.get("COMISSAO")
            if policy_col is None or premio_col is None or comissao_col is None:
                continue
            for row_index in range(header_row + 1, ws.max_row + 1):
                policy_id = str(ws.cell(row_index, policy_col).value or "").strip()
                if not policy_id:
                    continue
                parsed.append(
                    SegfyPolicyData(
                        policy_id=policy_id,
                        premio_total=_to_decimal(ws.cell(row_index, premio_col).value),
                        comissao=_to_decimal(ws.cell(row_index, comissao_col).value),
                    )
                )
        return parsed

    def _find_headers(self, ws) -> tuple[int | None, dict[str, int]]:
        aliases: dict[str, set[str]] = {
            "POLICY": {"POLICY", "POLICY ID", "APOLICE", "NUMERO APOLICE", "N APOLICE"},
            "PREMIO": {"PREMIO", "PREMIO TOTAL", "VALOR PREMIO"},
            "COMISSAO": {"COMISSAO", "VALOR COMISSAO"},
            "SEGURADO": {"SEGURADO", "SEGURADO(A)", "SEGURADA", "NOME", "CLIENTE", "SEGURADO A"},
            "SEGURADORA": {"SEGURADORA", "CIA", "COMPANHIA", "EMPRESA", "SEGURADORA CIA"},
            "VIG": {
                "VIG", "VIGENCIA", "VIG.", "VENCIMENTO", "DATA VIG", "VIG FIM",
                "VENCIMENTO APOLICE", "VIGENCIA FIM", "DATA VIGENCIA", "FIM VIGENCIA",
            },
            "STATUS_PGTO": {
                "STATUS PGTO", "STATUS PAGAMENTO", "PAGAMENTO", "PGTO", "STATUS",
                "SITUACAO PGTO", "SITUACAO PAGAMENTO",
            },
            "SINISTRO": {"SINISTRO", "SINISTRO ABERTO", "OCORRENCIA", "SINISTROS"},
            "ENDOSSO": {"ENDOSSO", "ENDOSSO ABERTO", "ENDOSSOS"},
            "ITEM": {"ITEM", "VEICULO", "ITEM SEGURADO", "PRODUTO", "BEM SEGURADO"},
            "MODELO": {"MODELO", "DESCRICAO", "DESCR", "DESCRICAO VEICULO", "MODELO VEICULO"},
            "PLACA": {"PLACA", "PLACA VEICULO", "PLACA DO VEICULO"},
        }
        for row_index in range(1, min(ws.max_row, 50) + 1):
            resolved: dict[str, int] = {}
            for col in range(1, ws.max_column + 1):
                normalized = _normalize(ws.cell(row_index, col).value)
                if normalized == "":
                    continue
                for logical_name, values in aliases.items():
                    if normalized in values and logical_name not in resolved:
                        resolved[logical_name] = col
            if {"POLICY", "PREMIO", "COMISSAO"}.issubset(set(resolved.keys())):
                return row_index, resolved
        return None, {}

    def _fetch_full_policies_from_export(self) -> list[PolicyRecord]:
        if self.export_xlsx_path is None or not self.export_xlsx_path.exists():
            return []

        today = date.today()
        workbook = load_workbook(self.export_xlsx_path, data_only=True)
        parsed: list[PolicyRecord] = []
        for ws in workbook.worksheets:
            header_row, header_map = self._find_headers(ws)
            if header_row is None:
                continue
            policy_col = header_map.get("POLICY")
            if policy_col is None:
                continue
            premio_col = header_map.get("PREMIO")
            comissao_col = header_map.get("COMISSAO")
            segurado_col = header_map.get("SEGURADO")
            seguradora_col = header_map.get("SEGURADORA")
            vig_col = header_map.get("VIG")
            status_pgto_col = header_map.get("STATUS_PGTO")
            sinistro_col = header_map.get("SINISTRO")
            endosso_col = header_map.get("ENDOSSO")
            item_col = header_map.get("ITEM")
            modelo_col = header_map.get("MODELO")
            placa_col = header_map.get("PLACA")

            for row_index in range(header_row + 1, ws.max_row + 1):
                policy_id = str(ws.cell(row_index, policy_col).value or "").strip()
                if not policy_id:
                    continue

                vig_raw = ws.cell(row_index, vig_col).value if vig_col else None
                vig = _to_date(vig_raw) if vig_raw is not None else today

                sinistro_raw = _normalize(ws.cell(row_index, sinistro_col).value) if sinistro_col else ""
                endosso_raw = _normalize(ws.cell(row_index, endosso_col).value) if endosso_col else ""
                sinistro_open = sinistro_raw in {"SIM", "S", "1", "ABERTO", "OPEN", "X"}
                endosso_open = endosso_raw in {"SIM", "S", "1", "ABERTO", "OPEN", "X"}

                parsed.append(PolicyRecord(
                    policy_id=policy_id,
                    insured_name=str(ws.cell(row_index, segurado_col).value or "").strip() if segurado_col else "",
                    insurer=str(ws.cell(row_index, seguradora_col).value or "").strip() if seguradora_col else "",
                    vig=vig or today,
                    status_pgto=str(ws.cell(row_index, status_pgto_col).value or "").strip() if status_pgto_col else "",
                    sinistro_open=sinistro_open,
                    endosso_open=endosso_open,
                    premio_total=_to_decimal(ws.cell(row_index, premio_col).value) if premio_col else Decimal("0"),
                    comissao=_to_decimal(ws.cell(row_index, comissao_col).value) if comissao_col else Decimal("0"),
                    vehicle_item=str(ws.cell(row_index, item_col).value or "").strip() if item_col else "",
                    vehicle_model=str(ws.cell(row_index, modelo_col).value or "").strip() if modelo_col else "",
                    vehicle_plate=str(ws.cell(row_index, placa_col).value or "").strip() if placa_col else "",
                ))
        return parsed
