from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
import re
import unicodedata

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from rpa_corretora.domain.models import CashflowEntry, ExpenseEntry, FollowupRecord, PolicyRecord


MONTH_SHEETS = {"JANEIRO", "FEVEREIRO", "MARCO", "ABRIL", "MAIO", "JUNHO", "JULHO", "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO", "DEZEMBRO"}
POLICY_HEADER_MIN = {"VIG", "SEGURADO(A)", "SEGURADORA"}
FOLLOWUP_HEADER_MIN = {"FASE", "STATUS", "NOVOS"}
RENDIMENTO_HEADER_MIN = {"DATA", "VALOR", "SEGURADORA", "ESPECIFICACAO"}
PLATE_PATTERNS = (
    re.compile(r"\b([A-Z]{3}[0-9][A-Z][0-9]{2})\b"),  # Mercosul
    re.compile(r"\b([A-Z]{3}[0-9]{4})\b"),  # Formato antigo
)


def _slug(value: str) -> str:
    cleaned = _normalize(value)
    cleaned = re.sub(r"[^A-Z0-9]+", "-", cleaned).strip("-")
    return cleaned or "SEM-ID"


def _extract_vehicle_info(item: str) -> tuple[str, str]:
    upper_item = _normalize(item)
    plate = ""
    for pattern in PLATE_PATTERNS:
        match = pattern.search(upper_item)
        if match is not None:
            plate = match.group(1)
            break

    model = item.strip()
    if plate:
        model = re.sub(re.escape(plate), "", upper_item, flags=re.IGNORECASE).replace("-", " ").strip()
        model = model or item.strip()
    return model, plate


def _normalize(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text == "":
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    return text.strip().upper()


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").replace("\t", " ").strip()
    return " ".join(text.split())


def _to_date(value: object) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _clean_text(value)
    if not text:
        return None
    for separator in ("/", "-"):
        parts = text.split(separator)
        if len(parts) == 3:
            try:
                if len(parts[0]) == 4:
                    return date(int(parts[0]), int(parts[1]), int(parts[2]))
                return date(int(parts[2]), int(parts[1]), int(parts[0]))
            except ValueError:
                pass
    return None


def _to_decimal(value: object) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = _clean_text(value).replace(".", "").replace(",", ".")
    if text == "":
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def _find_header_row(ws: Worksheet, required_headers: set[str], max_scan_row: int = 50) -> tuple[int, dict[str, int]] | None:
    scan_limit = min(max_scan_row, ws.max_row)
    for row_index in range(1, scan_limit + 1):
        header_map: dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            normalized = _normalize(ws.cell(row_index, col).value)
            if normalized:
                header_map[normalized] = col
        if required_headers.issubset(set(header_map.keys())):
            return row_index, header_map
    return None


def _has_content(values: list[object]) -> bool:
    return any(_clean_text(value) for value in values)


def _is_open_flag(value: object) -> bool:
    normalized = _normalize(value)
    if normalized == "":
        return False
    closed_tokens = {"CONCLUIDO", "CONCLUIDA", "FINALIZADO", "FINALIZADA", "RESOLVIDO", "RESOLVIDA", "NAO", "N/A"}
    return normalized not in closed_tokens


def _is_valid_followup_name(value: object) -> bool:
    text = _clean_text(value)
    if not text:
        return False
    normalized = _normalize(text)
    blocked_prefixes = (
        "TOTAL",
        "CENARIO",
        "FECHADOS",
        "NOVO",
        "RENOV",
        "N RENOV",
        "ENDOSSOS",
        "COPIE E COLE",
    )
    if normalized.startswith(blocked_prefixes):
        return False
    if normalized in {"-", "`"}:
        return False
    return True


class WorkbookSpreadsheetGateway:
    def __init__(
        self,
        seguros_pbseg_path: str | Path,
        acompanhamento_path: str | Path,
        fluxo_caixa_path: str | Path,
    ) -> None:
        self.seguros_pbseg_path = Path(seguros_pbseg_path)
        self.acompanhamento_path = Path(acompanhamento_path)
        self.fluxo_caixa_path = Path(fluxo_caixa_path)

    def _resolve_policy_id(
        self,
        *,
        ws: Worksheet,
        row_index: int,
        header_map: dict[str, int],
        insured_name: str,
        insurer: str,
        vig: date,
        duplicate_counter: dict[str, int],
    ) -> str:
        policy_number_columns = (
            "APOLICE",
            "APOLICE N",
            "N APOLICE",
            "NUMERO APOLICE",
            "N DA APOLICE",
            "APOLICE CONTRATO",
            "N DA APOLICE CONTRATO",
        )
        for header_name in policy_number_columns:
            col = header_map.get(header_name)
            if col is None:
                continue
            raw_policy = _clean_text(ws.cell(row_index, col).value)
            if raw_policy:
                return raw_policy

        base = f"{_slug(insured_name)[:20]}-{vig.strftime('%Y%m%d')}-{_slug(insurer)[:12]}"
        duplicate_counter[base] += 1
        suffix = duplicate_counter[base]
        if suffix > 1:
            return f"{base}-{suffix}"
        return base

    def load_policies(self) -> list[PolicyRecord]:
        workbook = load_workbook(self.seguros_pbseg_path, data_only=True)
        policies: list[PolicyRecord] = []
        duplicate_counter: dict[str, int] = defaultdict(int)

        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            header = _find_header_row(ws, POLICY_HEADER_MIN)
            if header is None:
                continue

            header_row, header_map = header
            segurado_col = header_map.get("SEGURADO(A)")
            seguradora_col = header_map.get("SEGURADORA")
            vig_col = header_map.get("VIG")
            sinistro_col = header_map.get("SINISTRO")
            endosso_col = header_map.get("ENDOSSO")
            status_pgto_col = header_map.get("STATUS PGTO")
            pt_col = header_map.get("PT")
            comissao_col = header_map.get("COMISSAO")
            item_col = header_map.get("ITEM")

            if segurado_col is None or seguradora_col is None or vig_col is None:
                continue

            for row_index in range(header_row + 1, ws.max_row + 1):
                segurado = _clean_text(ws.cell(row_index, segurado_col).value)
                seguradora = _clean_text(ws.cell(row_index, seguradora_col).value)
                vig = _to_date(ws.cell(row_index, vig_col).value)

                if not segurado and not seguradora and vig is None:
                    continue
                if not segurado or vig is None:
                    continue

                status_pgto = _clean_text(ws.cell(row_index, status_pgto_col).value) if status_pgto_col else ""
                sinistro_open = _is_open_flag(ws.cell(row_index, sinistro_col).value) if sinistro_col else False
                endosso_open = _is_open_flag(ws.cell(row_index, endosso_col).value) if endosso_col else False

                premio_total = _to_decimal(ws.cell(row_index, pt_col).value) if pt_col else Decimal("0")
                comissao = _to_decimal(ws.cell(row_index, comissao_col).value) if comissao_col else Decimal("0")
                item = _clean_text(ws.cell(row_index, item_col).value) if item_col else ""
                vehicle_model, vehicle_plate = _extract_vehicle_info(item)

                checked_value = ws.cell(row_index, 1).value
                renewal_started = bool(checked_value) if isinstance(checked_value, bool) else bool(_clean_text(checked_value))

                policy_id = self._resolve_policy_id(
                    ws=ws,
                    row_index=row_index,
                    header_map=header_map,
                    insured_name=segurado,
                    insurer=seguradora or "NAO INFORMADA",
                    vig=vig,
                    duplicate_counter=duplicate_counter,
                )

                policies.append(
                    PolicyRecord(
                        policy_id=policy_id,
                        insured_name=segurado,
                        insurer=seguradora or "NAO INFORMADA",
                        vig=vig,
                        renewal_kind="RENOVACAO_INTERNA",
                        renewal_started=renewal_started,
                        status_pgto=status_pgto,
                        sinistro_open=sinistro_open,
                        endosso_open=endosso_open,
                        premio_total=premio_total,
                        comissao=comissao,
                        vehicle_item=item,
                        vehicle_model=vehicle_model,
                        vehicle_plate=vehicle_plate,
                    )
                )

        return policies

    def load_followups(self, year: int) -> list[FollowupRecord]:
        _ = year
        workbook = load_workbook(self.acompanhamento_path, data_only=True)
        records: list[FollowupRecord] = []

        for sheet_name in workbook.sheetnames:
            normalized_sheet = _normalize(sheet_name)
            if normalized_sheet == "MASTER" or normalized_sheet not in MONTH_SHEETS:
                continue

            ws = workbook[sheet_name]
            header = _find_header_row(ws, FOLLOWUP_HEADER_MIN)
            if header is None:
                continue

            header_row, header_map = header
            internal_col = header_map.get("RENOVACOES INTERNAS")
            novos_col = header_map.get("NOVOS")
            if internal_col is None or novos_col is None:
                continue

            internal_vig_col = internal_col - 1
            internal_name_col = internal_col
            internal_fase_col = internal_col + 1
            internal_status_col = internal_col + 2

            novos_name_col = novos_col - 1
            novos_fase_col = novos_col + 1
            novos_status_col = novos_col + 2

            for row_index in range(header_row + 1, ws.max_row + 1):
                row_values = [ws.cell(row_index, c).value for c in range(1, ws.max_column + 1)]
                normalized_row = {_normalize(value) for value in row_values if _normalize(value)}
                if "ENDOSSOS" in normalized_row:
                    break

                internal_name_raw = ws.cell(row_index, internal_name_col).value
                novos_name_raw = ws.cell(row_index, novos_name_col).value

                if _is_valid_followup_name(internal_name_raw):
                    records.append(
                        FollowupRecord(
                            insured_name=_clean_text(internal_name_raw),
                            month=normalized_sheet,
                            fase=_clean_text(ws.cell(row_index, internal_fase_col).value),
                            status=_clean_text(ws.cell(row_index, internal_status_col).value),
                            renewal_kind="RENOVACAO_INTERNA",
                        )
                    )

                if _is_valid_followup_name(novos_name_raw):
                    records.append(
                        FollowupRecord(
                            insured_name=_clean_text(novos_name_raw),
                            month=normalized_sheet,
                            fase=_clean_text(ws.cell(row_index, novos_fase_col).value),
                            status=_clean_text(ws.cell(row_index, novos_status_col).value),
                            renewal_kind="NOVO",
                        )
                    )

                row_guard = [
                    ws.cell(row_index, internal_vig_col).value,
                    internal_name_raw,
                    ws.cell(row_index, internal_fase_col).value,
                    ws.cell(row_index, internal_status_col).value,
                    novos_name_raw,
                    ws.cell(row_index, novos_fase_col).value,
                    ws.cell(row_index, novos_status_col).value,
                ]
                if not _has_content(row_guard) and row_index > header_row + 10:
                    break

        return records

    def load_expenses(self, year: int, month: int) -> list[ExpenseEntry]:
        workbook = load_workbook(self.fluxo_caixa_path, data_only=True)
        ws = workbook["Gastos Mensais"]

        header = _find_header_row(ws, {"DATA", "DESCRICAO", "CATEGORIA", "VALOR (R$)"})
        if header is None:
            return []

        header_row, header_map = header
        data_col = header_map["DATA"]
        descricao_col = header_map["DESCRICAO"]
        categoria_col = header_map["CATEGORIA"]
        valor_col = header_map["VALOR (R$)"]

        expenses: list[ExpenseEntry] = []
        for row_index in range(header_row + 1, ws.max_row + 1):
            raw_date = ws.cell(row_index, data_col).value
            raw_value = ws.cell(row_index, valor_col).value
            if raw_date in (None, "") and raw_value in (None, ""):
                continue

            entry_date = _to_date(raw_date)
            if entry_date is None:
                continue
            if entry_date.year != year or entry_date.month != month:
                continue

            value = _to_decimal(raw_value)
            expenses.append(
                ExpenseEntry(
                    date=entry_date,
                    value=value,
                    description=_clean_text(ws.cell(row_index, descricao_col).value),
                    category=_clean_text(ws.cell(row_index, categoria_col).value),
                )
            )

        return expenses

    def append_cashflow_entries(self, entries: list[CashflowEntry]) -> None:
        if not entries:
            return

        workbook = load_workbook(self.fluxo_caixa_path)
        ws = workbook["RENDIMENTO"]

        header = _find_header_row(ws, RENDIMENTO_HEADER_MIN)
        if header is None:
            raise ValueError("Nao foi possivel localizar cabecalho DATA/VALOR/SEGURADORA/ESPECIFICACAO na aba RENDIMENTO")

        header_row, header_map = header
        data_col = header_map["DATA"]
        valor_col = header_map["VALOR"]
        seguradora_col = header_map["SEGURADORA"]
        especificacao_col = header_map["ESPECIFICACAO"]

        write_row = header_row + 1
        while True:
            row_has_data = any(
                ws.cell(write_row, col).value not in (None, "")
                for col in (data_col, valor_col, seguradora_col, especificacao_col)
            )
            if not row_has_data:
                break
            write_row += 1

        for entry in entries:
            ws.cell(write_row, data_col).value = entry.date
            ws.cell(write_row, valor_col).value = float(entry.value)
            ws.cell(write_row, seguradora_col).value = entry.insurer
            ws.cell(write_row, especificacao_col).value = entry.specification
            write_row += 1

        workbook.save(self.fluxo_caixa_path)

    def append_expense_entries(self, entries: list[ExpenseEntry]) -> None:
        if not entries:
            return

        workbook = load_workbook(self.fluxo_caixa_path)
        ws = workbook["Gastos Mensais"]

        header = _find_header_row(ws, {"DATA", "DESCRICAO", "CATEGORIA", "VALOR (R$)"})
        if header is None:
            raise ValueError("Nao foi possivel localizar cabecalho DATA/DESCRICAO/CATEGORIA/VALOR (R$) na aba Gastos Mensais")

        header_row, header_map = header
        data_col = header_map["DATA"]
        descricao_col = header_map["DESCRICAO"]
        categoria_col = header_map["CATEGORIA"]
        valor_col = header_map["VALOR (R$)"]

        write_row = header_row + 1
        while True:
            row_has_data = any(
                ws.cell(write_row, col).value not in (None, "")
                for col in (data_col, descricao_col, categoria_col, valor_col)
            )
            if not row_has_data:
                break
            write_row += 1

        for entry in entries:
            ws.cell(write_row, data_col).value = entry.date
            ws.cell(write_row, descricao_col).value = entry.description
            ws.cell(write_row, categoria_col).value = entry.category
            ws.cell(write_row, valor_col).value = float(entry.value)
            write_row += 1

        workbook.save(self.fluxo_caixa_path)

    def validate_expense_summary(self, year: int, month: int) -> list[str]:
        expenses = self.load_expenses(year, month)
        computed_by_category: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
        for expense in expenses:
            key = _normalize(expense.category)
            if key == "":
                key = "SEM CATEGORIA"
            computed_by_category[key] += expense.value

        workbook = load_workbook(self.fluxo_caixa_path, data_only=True)
        ws = workbook["Resumo De Gastos"]

        header = _find_header_row(ws, {"CATEGORIA"})
        if header is None:
            return ["Resumo de gastos: cabecalho de categoria nao encontrado."]
        header_row, header_map = header
        category_col = header_map["CATEGORIA"]
        total_col = header_map.get("TOTAL (SUMIF - EN)")
        if total_col is None:
            total_col = header_map.get("TOTAL (SOMASE - PT-BR)")
        if total_col is None:
            return ["Resumo de gastos: coluna de total nao encontrada."]

        issues: list[str] = []
        for row_index in range(header_row + 1, ws.max_row + 1):
            category_raw = _clean_text(ws.cell(row_index, category_col).value)
            if not category_raw:
                continue
            normalized_category = _normalize(category_raw)
            expected = computed_by_category.get(normalized_category, Decimal("0"))
            raw_total = ws.cell(row_index, total_col).value
            observed = _to_decimal(raw_total)

            # Quando o Excel nao recalculou formulas (openpyxl), o valor pode vir vazio.
            if raw_total in (None, ""):
                continue

            if abs(observed - expected) > Decimal("0.01"):
                issues.append(
                    (
                        f"Resumo divergente em '{category_raw}': "
                        f"esperado={expected:.2f} observado={observed:.2f}"
                    )
                )

        return issues
